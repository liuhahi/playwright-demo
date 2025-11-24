from __future__ import annotations
import os
import io
import json
import base64
import time
import argparse
from typing import List, Dict, Any, Optional
from typing_extensions import Literal
from dataclasses import dataclass, asdict
import hashlib
from datetime import datetime, timezone

from openai import OpenAI
import re
import requests
import textwrap
from urllib.parse import urljoin, urlparse
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import yaml
from pydantic import BaseModel

WRITE_LIVE_GRAPH=True

# NOTE: Feature nodes are LLM-gated. Only clickables approved by llm_filter_clickables_with_prior (llm_keep=True)
# will be written to the graph, preventing noise like numbers, file names, or table headers.

# ---------------- Dataclasses ----------------
class FeatureItem(BaseModel):
    id: str
    name: str
    category: str
    actions: list[str]
    page: str
    evidence: str
    locator: str = ""
    confidence: int

class FeaturesSchema(BaseModel):
    features: list[FeatureItem]

class ReconciledItem(BaseModel):
    id: str
    name: str
    category: str
    actions: list[str]
    page: str
    evidence: str
    locator: str = ""
    confidence: int
    status: str = "both"
    matched_with: list[str] = []

class ReconciledSchema(BaseModel):
    features: list[ReconciledItem]

class UserStoryItem(BaseModel):
    id: str
    actor: str
    title: str
    story: str
    acceptance_criteria: list[str]
    related_features: list[str]
    priority: str

class UserStoriesSchema(BaseModel):
    stories: list[UserStoryItem]

# --- Prior Features Pydantic Schemas ---
class PriorFeatureItem(BaseModel):
    id: str
    name: str
    actions: list[str]
    role: str  # e.g., "user", "admin", "viewer"

class PriorFeaturesSchema(BaseModel):
    prior_features: list[PriorFeatureItem]

class NewFeatureItem(BaseModel):
    name: str
    actions: list[str]
    role: str

class ActionFeatureMapping(BaseModel):
    prior_ids: list[str]
    new_features: list[NewFeatureItem]

# Pydantic schema for next-action choice
class ActionChoice(BaseModel):
    action: Literal["click", "back", "stop"]
    target: Optional[str] = None
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from PIL import Image

# --- Neo4j/Working Memory graph helpers ---
from working_memory_importer import WebsiteWMImporter


# --- Working Memory graph singleton and toggle ---
WM_IMPORTER: WebsiteWMImporter | None = None
WRITE_LIVE_GRAPH: bool = True
PRIOR_FEATURES: list[PriorFeature] | None = None  # populated in agent_explore_once for LLM mapping
PRIOR_PATH: str | None = None  # path to prior_features.json for incremental updates

"""
This script analyzes a website using:
  1) HTML files you manually saved
  2) Screenshots (images) of pages

Pipeline:
  - Step 1: Extract features from HTML (LLM on DOM text only)
  - Step 2: Extract features from screenshots (LLM on image content only)
  - Step 3: Cross-check & reconcile (merge duplicates, mark source: both/html_only/image_only)
  - Step 4: Generate deduplicated User Stories (with Acceptance Criteria & Priority)
  - Export: Excel for features & stories + JSON dump

Usage:
  export OPENAI_API_KEY=sk-...
  python demo.py --html page1.html page2.html --shots ./screenshots --out ./out

Dependencies:
  pip install openai bs4 lxml pillow pandas openpyxl
"""

# ===================== Config =====================
# ===================== Config =====================
MODEL_FEATURE = "gpt-4.1-mini"   # Reasoning model for structure/merging (faster)
MODEL_VISION  = "gpt-4.1"    # Same model is fine; supports image inputs (faster)
MAX_RETRIES   = 3
REQ_TIMEOUT   = 120
HTML_CHUNK_CHARS = 15000       # Per-chunk size for long HTML (smaller chunks for speed)
IMG_MAX_SIDE  = 1600           # Downscale long side before sending
DOC_CRAWL_MAX_PAGES = 60      # Max number of documentation pages to crawl per docs root
# ==================================================

# ====== Agent (one-round) exploration config ======
AGENT_MAX_STEPS_PER_FEATURE = 12
AGENT_WAIT_NETWORK_IDLE_MS = 1200
AGENT_CLICK_TIMEOUT_MS = 5000
AGENT_SAFE_BUTTONS = {"Reports", "History"}  # Avoid mutating ops by default (e.g., Scan/Settings)

client = OpenAI()

# ---------------- Dataclasses ----------------
@dataclass
class Feature:
    id: str
    name: str
    category: str           # e.g., navigation, dashboard, list, detail, dialog, settings, notification
    actions: List[str]      # e.g., ["click","expand","filter","sort","search","paginate","trigger_scan"]
    page: str               # e.g., "Home", "Projects"
    evidence: str           # short snippet or hint
    source: str             # "html" | "image" | "merged"
    confidence: int         # 1..5
    locator: str = ""       # optional: selector/label/visible text

@dataclass
class ReconciledFeature(Feature):
    status: str = "both"    # "both" | "html_only" | "image_only"
    matched_with: List[str] = None  # ids it merged with

@dataclass
class UserStory:
    id: str
    actor: str
    title: str              # Verb + Noun
    story: str              # As a <actor>, I want to <goal>, so that <benefit>.
    acceptance_criteria: List[str]
    related_features: List[str]
    priority: str           # High | Medium | Low

# --- Prior Feature dataclass ---
@dataclass
class PriorFeature:
    id: str
    name: str
    actions: List[str]
    role: str


# --------------- Helpers -----------------

# -------- Prior (save/load) --------
def save_prior_features(prior: List[PriorFeature], path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump([
            {"id": x.id, "name": x.name, "actions": x.actions, "role": x.role}
            for x in prior
        ], f, ensure_ascii=False, indent=2)

def load_prior_features(path: str) -> List[PriorFeature]:
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)
    prior: List[PriorFeature] = []
    for it in raw:
        prior.append(PriorFeature(
            id=it.get("id", ""),
            name=it.get("name", ""),
            actions=it.get("actions", []) or [],
            role=it.get("role", "user"),
        ))
    return prior

def retrying(callable_fn):
    def _wrap(*args, **kwargs):
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                return callable_fn(*args, **kwargs)
            except Exception as e:
                if attempt == MAX_RETRIES:
                    raise
                time.sleep(6 * attempt)
        return None
    return _wrap


def read_html_as_chunks(path: str) -> List[str]:
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        raw = f.read()
    # Strip scripts/styles to reduce noise
    soup = BeautifulSoup(raw, "lxml")
    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()
    text = str(soup)
    # Chunked to control token usage
    return [text[i:i + HTML_CHUNK_CHARS] for i in range(0, len(text), HTML_CHUNK_CHARS)] or [text]


def image_to_base64(path: str) -> (str, str):
    img = Image.open(path)
    w, h = img.size
    mx = max(w, h)
    if mx > IMG_MAX_SIDE:
        scale = IMG_MAX_SIDE / mx
        img = img.resize((int(w * scale), int(h * scale)))
    buf = io.BytesIO()
    fmt = "PNG" if path.lower().endswith(".png") else "JPEG"
    img.save(buf, format=fmt)
    mime = "image/png" if fmt == "PNG" else "image/jpeg"
    return base64.b64encode(buf.getvalue()).decode("utf-8"), mime


# --------------- Docs → Prior Features (one-round agent) ---------------

def fetch_docs_text(urls: List[str]) -> str:
    """
    Fetch documentation pages and keep mostly natural-language text.
    We strip scripts/styles and also code/pre blocks to avoid noisy code samples.
    """
    chunks = []
    headers = {"User-Agent": "web-reverse-agent/1.0"}
    for u in urls:
        try:
            r = requests.get(u, headers=headers, timeout=20)
            r.raise_for_status()
            html = r.text
            # Use BeautifulSoup to remove non-natural-language parts
            soup = BeautifulSoup(html, "lxml")
            # Drop scripts, styles, noscript and code/pre blocks (code samples)
            for tag in soup(["script", "style", "noscript", "code", "pre"]):
                tag.decompose()
            # Extract mostly natural language text
            text = soup.get_text(separator=" ", strip=True)
            chunks.append(f"\n[DOC] {u}\n{text[:100000]}")  # cap to 100k chars per doc
        except Exception as e:
            chunks.append(f"\n[DOC-ERROR] {u}: {e}")
    return "\n\n".join(chunks)

# --- Crawl all pages under the given docs root(s) ---
def crawl_docs_text(root_urls: List[str], max_pages: int = DOC_CRAWL_MAX_PAGES) -> str:
    """
    Crawl documentation starting from the given root URLs.
    - Stays within the same host and path prefix as each root URL.
    - Breadth-first crawl up to max_pages pages in total.
    - Returns a big concatenated text string for all fetched pages.
    """
    headers = {"User-Agent": "web-reverse-agent/1.0"}
    seen: set[str] = set()
    queue: List[str] = []
    chunks: List[str] = []

    # Normalize roots and seed the queue
    normalized_roots: List[tuple[str, str, str]] = []  # (root_url, netloc, path_prefix)
    for ru in root_urls:
        ru = ru.strip()
        if not ru:
            continue
        parsed = urlparse(ru)
        if not parsed.scheme.startswith("http"):
            continue
        path_prefix = parsed.path.rstrip("/") or "/"
        normalized_roots.append((ru, parsed.netloc, path_prefix))
        if ru not in queue:
            queue.append(ru)

    def _belongs_to_roots(url: str) -> bool:
        try:
            p = urlparse(url)
        except Exception:
            return False
        if p.scheme not in ("http", "https"):
            return False
        for _, netloc, path_prefix in normalized_roots:
            if p.netloc == netloc and (p.path or "/").startswith(path_prefix):
                return True
        return False

    while queue and len(seen) < max_pages:
        u = queue.pop(0)
        if u in seen:
            continue
        seen.add(u)
        # Log which page is being crawled
        try:
            print(f"[DOC-CRAWL] Fetching {u} ({len(seen)}/{max_pages})")
        except Exception:
            # Avoid breaking the crawl on encoding issues
            pass
        try:
            r = requests.get(u, headers=headers, timeout=20)
            r.raise_for_status()
            html = r.text
            # Parse HTML and strip scripts/styles/code/pre/noscript to reduce noise
            soup = BeautifulSoup(html, "lxml")
            # Remove scripts/styles/noscript and code/pre blocks to focus on natural language docs
            for tag in soup(["script", "style", "noscript", "code", "pre"]):
                tag.decompose()
            # Use text content for LLM (cleaner than raw HTML, less code noise)
            text = soup.get_text(separator=" ", strip=True)
            chunks.append(f"\n[DOC] {u}\n{text[:100000]}")  # cap to 100k chars per page

            # Discover more links within the same docs domain/prefix
            for a in soup.find_all("a", href=True):
                href = (a.get("href") or "").strip()
                if not href:
                    continue
                full = urljoin(u, href)
                if full in seen or full in queue:
                    continue
                if not _belongs_to_roots(full):
                    continue
                queue.append(full)
        except Exception as e:
            chunks.append(f"\n[DOC-ERROR] {u}: {e}")

    return "\n\n".join(chunks)

# --------- Prior features from docs only ---------
@retrying
def llm_prior_features_from_docs(doc_text: str) -> List[PriorFeature]:
    """Produce a PRIOR feature list directly from online docs only."""
    prompt = (
        "You are acting as a product-UX analyst agent.\n"
        "From the following documentation text (web pages), infer LIKELY user-facing FEATURES of the web app.\n"
        "For each feature, return: id (kebab-case), name (visible UI label), actions (list of user operations such as click/search/filter), and role (who performs the action, e.g., user/admin/viewer).\n"
        "Do not include backend-only APIs."
    )
    resp = client.responses.parse(
        model=MODEL_FEATURE,
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": prompt},
                {"type": "input_text", "text": doc_text[:200000]},
            ]
        }],
        text_format=PriorFeaturesSchema,
        timeout=REQ_TIMEOUT,
    )
    data: PriorFeaturesSchema = resp.output_parsed
    out: List[PriorFeature] = []
    # Assign sequential, index-based IDs (feature1, feature2, ...) to prior features
    for idx, f in enumerate(data.prior_features, start=1):
        out.append(PriorFeature(
            id=f"feature{idx}",
            name=f.name,
            actions=f.actions,
            role=f.role,
        ))
    return out

# --------------- LLM Calls -----------------
# --------------- One-Round Agent: Playwright Exploration ---------------

def _normalize_path(pathname: str) -> str:
    if not pathname:
        return "/"
    p = pathname.split("?")[0]
    p = re.sub(r"/org/[0-9a-fA-F-]{6,}", "/org/:orgId", p)
    p = re.sub(r"/[0-9]+(?=/|$)", "/:id", p)
    p = re.sub(r"/[0-9a-f]{8}-[0-9a-f]{4}-[1-5][0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12}(?=/|$)", "/:id", p, flags=re.I)
    p = re.sub(r"/[0-9a-f]{16,}(?=/|$)", "/:id", p, flags=re.I)
    return p.rstrip("/") or "/"

def _record_route(routes: set, url: str):
    pass

def _utcnow_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

def _sha256(text: str) -> str:
    try:
        return hashlib.sha256(text.encode("utf-8", errors="ignore")).hexdigest()
    except Exception:
        return ""

def _page_route_and_title(pg) -> tuple[str, str]:
    try:
        rt = _normalize_path(urlparse(pg.url).path)
    except Exception:
        rt = "/"
    try:
        ttl = pg.title() or ""
    except Exception:
        ttl = ""
    return rt, ttl

# Helper: derive a more distinctive page title from route/URL when the raw title is generic/empty
def _derive_distinct_title(url: str, route: str, raw_title: str) -> str:
    """Make page titles distinct: prefer sanitized <title>, else derive from last meaningful route segment.
    Examples:
      /projects -> "Projects"
      /projects/:id -> "Project Detail"
      /org/:orgId/dashboard -> "Dashboard"
    """
    t = _sanitize_page_title(raw_title)
    if t:
        return t
    # fallback: derive from route
    segs = [s for s in route.split('/') if s and not s.startswith(':')]
    if not segs:
        # try hostname
        try:
            host = urlparse(url).hostname or ""
        except Exception:
            host = ""
        return host.split('.')[-2].title() if host else "Page"
    last = segs[-1]
    # normalize
    last = last.replace('-', ' ').replace('_', ' ').strip()
    if last:
        # special-cases
        if last in {"projects", "project"} and ":id" in route:
            return "Project Detail"
        if last in {"org", "organization"}:
            return "Organization"
        return last.title()
    return "Page"

# Helper to sanitize page titles by removing generic site names and common patterns
def _sanitize_page_title(title: str) -> str:
    """Remove generic site names. Return empty string if only site name remains."""
    t = (title or "").strip()
    if not t:
        return ""
    # If title is just the site name, drop it
    if t.lower() in {"scantist", "scantist devsecops", "scantist devsec ops"}:
        return ""
    # Remove common prefix/suffix decorations with the site name
    import re as _re
    t = _re.sub(r"\s*[-|•·–—]\s*scantist(?:\s+devsec\s*ops)?\s*$", "", t, flags=_re.I)
    t = _re.sub(r"^\s*scantist(?:\s+devsec\s*ops)?\s*[-|•·–—]\s*", "", t, flags=_re.I)
    # If after stripping it becomes empty, return empty
    return t.strip()

def _wm_connect(uri: str, user: str, password: str):
    global WM_IMPORTER
    if WM_IMPORTER is None:
        imp = WebsiteWMImporter(uri, user, password)
        imp.connect()
        imp.ensure_schema()
        WM_IMPORTER = imp

def _wm_close():
    global WM_IMPORTER
    if WM_IMPORTER is not None:
        WM_IMPORTER.close()
        WM_IMPORTER = None

def _wm_upsert_page_from_playwright(pg):
    if not WRITE_LIVE_GRAPH or WM_IMPORTER is None:
        return
    route, raw_title = _page_route_and_title(pg)
    title = _derive_distinct_title(pg.url, route, raw_title)
    html = ""
    try:
        html = pg.content()
    except Exception:
        pass
    html_hash = _sha256(html) if html else None
    WM_IMPORTER.upsert_page({
        "url": pg.url,
        "route": pg.url,  # use full URL as unique route so each URL becomes a distinct node
        "title": title or None,
        "status": None,
        "html_hash": html_hash,
        "last_crawled_at": _utcnow_iso(),
    })

def _wm_upsert_feature_for_clickable(page_url: str, clickable: dict):
    if not WRITE_LIVE_GRAPH or WM_IMPORTER is None:
        return
    # Only allow features that passed LLM gating
    if not clickable.get("llm_keep"):
        return
    resolved = (clickable.get("canonical_name") or clickable.get("name") or "").strip()
    kind = clickable.get("category") or clickable.get("kind") or "generic"
    selector = f"text={resolved}" if resolved else "[data-auto]"
    WM_IMPORTER.upsert_feature({
        "page_url": page_url,
        "selector": selector,
        "kind": kind,
        "name": resolved,
        "text": resolved,
    })
# -------- Rank clickables for fallback (Vision + ARIA, no URL/HTML) --------
@retrying
def llm_filter_clickables_with_prior(clickables: List[dict], current_feature: PriorFeature, aria_yaml_text: str, screenshot_b64: str) -> List[dict]:
    """Ask the model which clickables are relevant to the current feature. Returns subset with llm_keep flags."""
    options = sorted({(c.get("name") or "").strip() for c in clickables if c.get("name")})
    aria_excerpt = aria_yaml_text[:15000] if aria_yaml_text else ""
    # Describe the single feature we are currently exploring
    feat_desc = {
        "id": current_feature.id,
        "name": current_feature.name,
        "role": current_feature.role,
        "actions": current_feature.actions,
    }
    system = (
        "You select UI elements for a web-exploration agent that is currently exploring ONE specific feature of the app.\n"
        "Your goal is to KEEP clickables that are likely RELEVANT steps or entry points for this target feature, and DROP those that are clearly unrelated.\n"
        "Be inclusive for anything that might help the agent navigate or use this feature."
    )
    instructions = (
        "Target feature (for context):\n"
        + json.dumps(feat_desc, ensure_ascii=False)
        + "\n\n"
        "You are given the list of clickable NAMES on the current page.\n"
        "For each clickable NAME, decide whether it could be a useful step (directly or indirectly) toward using or reaching this target feature.\n"
        "A clickable is considered relevant if clicking it might reasonably navigate closer to the feature, open a related section/tab, apply a filter, or drill down into data that belongs to this feature—even if it does NOT immediately open the final screen for this feature.\n"
        "Return a STRICT JSON list where each item corresponds to one input name:\n"
        "[{\"name\":\"...\",\"keep\":true|false,\"canonical_name\":\"<optional, if you want to normalize>\",\"category\":\"<optional coarse category>\"}]\n"
        "Guidelines:\n"
        "- Mark keep=true for elements that are likely useful steps on a path toward this feature: navigation into its area, tabs/filters/lists/details that look related, or links that could lead to relevant pages.\n"
        "- Mark keep=false only for items that are clearly unrelated noise for this feature: unrelated settings, language toggles, pure numbers, dates, obvious counters, or modules that clearly serve a different purpose.\n"
        "- When unsure whether an element might help eventually reach or use this feature, prefer keep=true so the agent can still try it as a potential step.\n"
    )
    content = [
        {"type": "input_text", "text": system},
        {"type": "input_text", "text": instructions},
        {"type": "input_text", "text": "Clickables:\n" + json.dumps(options, ensure_ascii=False)},
        {"type": "input_text", "text": "[ARIA EXCERPT]\n" + aria_excerpt},
    ]
    if screenshot_b64:
        content.insert(2, {"type": "input_image", "image_url": f"data:image/png;base64,{screenshot_b64}"})
    try:
        resp = client.responses.create(model=MODEL_VISION, input=[{"role": "user", "content": content}], timeout=REQ_TIMEOUT)
        text = resp.output_text
        decisions = json.loads(text)
    except Exception:
        decisions = []
    by_name = {(d.get("name") or "").strip().lower(): d for d in decisions if isinstance(d, dict)}
    kept: list[dict] = []

    for c in clickables:
        nm = (c.get("name") or "").strip()
        if not nm:
            continue
        d = by_name.get(nm.lower())
        cc = dict(c)
        # 默认行为：如果 LLM 没有给出明确决策，则保留该 clickable（视为潜在相关）
        if not d:
            cc["llm_keep"] = True
            cc["canonical_name"] = nm
            cc["category"] = c.get("kind") or "generic"
            kept.append(cc)
            continue
        # LLM 给出明确 keep=false：视为与当前 feature 无关，丢弃
        if d.get("keep") is False:
            continue
        # keep=true：保留并应用规范化信息
        cc["llm_keep"] = True
        canon = d.get("canonical_name") or nm
        cc["canonical_name"] = canon
        cc["category"] = d.get("category") or c.get("kind") or "generic"
        kept.append(cc)

    # 若解析失败或模型输出为空，fallback：保留所有原始 clickables
    if not decisions and not kept:
        fallback: list[dict] = []
        for c in clickables:
            nm = (c.get("name") or "").strip()
            if not nm:
                continue
            cc = dict(c)
            cc["llm_keep"] = True
            cc["canonical_name"] = nm
            cc["category"] = c.get("kind") or "generic"
            fallback.append(cc)
        return fallback

    return kept

def _wm_link_page_contains_features(page_url: str, clickables: list[dict]):
    if not WRITE_LIVE_GRAPH or WM_IMPORTER is None:
        return
    for c in clickables:
        _wm_upsert_feature_for_clickable(page_url, c)

def _wm_feature_to_page(src_page_url: str, clickable: dict, dst_page_url: str):
    """
    Deprecated: we no longer create per-clickable Feature nodes like
    '<page_url>|text=...'. Features are now represented only by PRIOR/LLM
    Feature nodes linked via ACHIEVE from Action nodes.
    """
    return

def _wm_page_to_page(src_url: str, dst_url: str, via: str = "click", anchor_text: str | None = None):
    """
    Deprecated: we no longer record Page↔Page relationships in the graph.
    Page–Page relationships can be reconstructed later via Page–Action–Page paths.
    """
    return

def _wm_calls_api_from_page(page_url: str, apis: dict):
    """
    Previously recorded Page ↔ APIEndpoint relationships in the graph.
    Now we keep API call info only in JSON artifacts; no graph edges are written.
    """
    return

# --- Action node helpers ---
def _wm_create_action_node(feature_id: str, page_url: str, action_type: str, name: str, description: str | None) -> str | None:
    """
    Create an Action node representing one LLM operation on a page, and link:
      (Page)-[:CALL]->(Action)

    The Action node ID is deterministic based on (feature_id, page_url, action_type, name)
    so that repeating the same action does not create duplicate nodes.
    """
    if not WRITE_LIVE_GRAPH or WM_IMPORTER is None:
        return None
    # Deterministic ID: same feature + page + action_type + name → same Action node
    normalized_name = (name or "").strip().lower()
    raw = f"{feature_id}|{page_url}|{action_type}|{normalized_name}"
    action_id = _sha256(raw)
    created_at = _utcnow_iso()
    try:
        driver = WM_IMPORTER._graph._driver
        with driver.session() as s:
            s.run(
                """
                MERGE (p:Page {url: $page_url})
                MERGE (a:Action {action_id: $action_id})
                ON CREATE SET
                    a.type = $action_type,
                    a.name = $name,
                    a.description = $description,
                    a.feature_id = $feature_id,
                    a.created_at = $created_at
                ON MATCH SET
                    a.type = coalesce($action_type, a.type),
                    a.name = coalesce($name, a.name),
                    a.description = coalesce($description, a.description),
                    a.feature_id = coalesce($feature_id, a.feature_id),
                    a.created_at = coalesce(a.created_at, $created_at)
                MERGE (p)-[:CALL]->(a)
                """,
                page_url=page_url,
                action_id=action_id,
                action_type=action_type,
                name=name,
                description=description,
                feature_id=feature_id,
                created_at=created_at,
            )
    except Exception:
        # Do not break the agent on graph errors
        return None
    return action_id


def _wm_link_action_next(prev_action_id: str | None, next_action_id: str | None):
    """Link two Action nodes with NEXT to indicate a multi-step sequence."""
    if not WRITE_LIVE_GRAPH or WM_IMPORTER is None:
        return
    if not prev_action_id or not next_action_id or prev_action_id == next_action_id:
        return
    try:
        driver = WM_IMPORTER._graph._driver
        with driver.session() as s:
            s.run(
                """
                MATCH (a1:Action {action_id: $prev}), (a2:Action {action_id: $next})
                MERGE (a1)-[:NEXT]->(a2)
                """,
                prev=prev_action_id,
                next=next_action_id,
            )
    except Exception:
        pass


def _wm_action_forward_page(action_id: str | None, dst_page_url: str):
    """Link an Action to the Page it navigates to via FORWARD."""
    if not WRITE_LIVE_GRAPH or WM_IMPORTER is None:
        return
    if not action_id or not dst_page_url:
        return
    try:
        driver = WM_IMPORTER._graph._driver
        with driver.session() as s:
            s.run(
                """
                MATCH (a:Action {action_id: $action_id})
                MERGE (p:Page {url: $dst_url})
                MERGE (a)-[:FORWARD]->(p)
                """,
                action_id=action_id,
                dst_url=dst_page_url,
            )
    except Exception:
        pass

@retrying
def llm_map_action_to_features(action_context: dict, prior_features: list[PriorFeature]) -> ActionFeatureMapping:
    """
    用 GPT 决定这个 Action 关联哪些 PRIOR features，以及是否需要创建新的非 prior features。
    """
    prior_list = [
        {"id": f.id, "name": f.name, "actions": f.actions, "role": f.role}
        for f in prior_features
    ][:60]

    prompt = (
        "You are a product-UX analyst working with a graph of Actions and Features.\n"
        "Given a single user Action (with its description and page context) and a list of PRIOR features,\n"
        "decide which prior features this action most likely ACHIEVES, and whether it implies any NEW\n"
        "non-prior user-facing features that are not covered by the prior list.\n\n"
        "Output JSON with two fields:\n"
        "  - prior_ids: list of feature ids from the prior list that this action achieves (can be empty or multiple).\n"
        "  - new_features: list of new features (if any), each with name, actions, and role.\n"
        "An action can achieve multiple related features. Keep new feature names short and Title-Case."
    )

    system_ctx = {
        "action": action_context,
        "prior_features": prior_list,
    }

    resp = client.responses.parse(
        model=MODEL_FEATURE,
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": prompt},
                {"type": "input_text", "text": json.dumps(system_ctx, ensure_ascii=False)},
            ]
        }],
        text_format=ActionFeatureMapping,
        timeout=REQ_TIMEOUT,
    )
    return resp.output_parsed

def _wm_action_achieve_feature(action_id: str | None, feature_id: str | None):
    """
    Link an Action to one or more Feature nodes it achieves via ACHIEVE.

    行为：
      1) 如果全局 PRIOR_FEATURES 存在，则调用 GPT 决定：
         - 这个 action 对哪些 prior features（id）有贡献；
         - 是否需要创建新的非 prior features。
      2) 所有 Feature 节点统一属性：
         feature_id, name, actions, role, created_at, last_seen_at, is_prior。
      3) 从这个 Action 连 ACHIEVE 到所有相关 Feature（可能多条边）。

    兼容：
      - 如果没有 PRIOR_FEATURES 或 GPT 出错，则退化为：只用传入的 feature_id 建一条边。
    """
    global PRIOR_FEATURES, PRIOR_PATH

    if not WRITE_LIVE_GRAPH or WM_IMPORTER is None:
        return
    if not action_id:
        return

    try:
        driver = WM_IMPORTER._graph._driver

        # 从图里取出 action 的基本信息，作为 LLM 的上下文
        with driver.session() as s:
            rec = s.run(
                """
                MATCH (a:Action {action_id: $action_id})
                OPTIONAL MATCH (a)-[:FORWARD]->(p:Page)
                RETURN a.name AS action_name,
                       a.type AS action_type,
                       a.description AS action_desc,
                       p.url AS page_url,
                       p.title AS page_title
                LIMIT 1
                """,
                action_id=action_id,
            ).single()

        action_ctx = {
            "name": rec["action_name"] if rec else None,
            "type": rec["action_type"] if rec else None,
            "description": rec["action_desc"] if rec else None,
            "page_url": rec["page_url"] if rec else None,
            "page_title": rec["page_title"] if rec else None,
            "fallback_feature_id": feature_id,
        }

        # 1) 调 GPT：决定 prior_ids + new_features
        mapped_prior_ids: list[str] = []
        new_features: list[NewFeatureItem] = []

        global PRIOR_FEATURES
        if PRIOR_FEATURES:
            try:
                mapping = llm_map_action_to_features(action_ctx, PRIOR_FEATURES)
                mapped_prior_ids = list(mapping.prior_ids or [])
                new_features = list(mapping.new_features or [])
            except Exception:
                mapped_prior_ids = []
                new_features = []

        # 退化：如果 GPT / PRIOR_FEATURES 不可用，就至少连到调用方提供的 feature_id
        if not mapped_prior_ids and feature_id:
            mapped_prior_ids = [feature_id]

        # 把 prior 写成 id -> meta 的 dict 方便查属性
        prior_by_id: dict[str, PriorFeature] = {}
        if PRIOR_FEATURES:
            for pf in PRIOR_FEATURES:
                prior_by_id[pf.id] = pf

        now_ms = int(datetime.now(timezone.utc).timestamp() * 1000)
        all_feature_ids: list[str] = []

        with driver.session() as s:
            # 2a) 确保所有 prior features 节点存在，并统一属性 + is_prior=true
            for fid in mapped_prior_ids:
                meta = prior_by_id.get(fid)
                name = meta.name if meta else fid
                actions = meta.actions if meta else []
                role = meta.role if meta else "user"

                s.run(
                    """
                    MERGE (ft:Feature {feature_id: $feature_id})
                    ON CREATE SET
                        ft.name = $name,
                        ft.actions = $actions,
                        ft.role = $role,
                        ft.created_at = $created_at,
                        ft.last_seen_at = $last_seen_at,
                        ft.is_prior = true
                    ON MATCH SET
                        ft.name = coalesce($name, ft.name),
                        ft.actions = coalesce($actions, ft.actions),
                        ft.role = coalesce($role, ft.role),
                        ft.last_seen_at = $last_seen_at,
                        ft.is_prior = true
                    """,
                    feature_id=fid,
                    name=name,
                    actions=actions,
                    role=role,
                    created_at=now_ms,
                    last_seen_at=now_ms,
                )
                all_feature_ids.append(fid)

            # 2b) 为 GPT 新生成的非 prior features 创建节点（is_prior=false），并增量写回 prior_features.json
            for nf in new_features:
                base_id = _slugify(nf.name)
                auto_id = f"auto-{base_id}" if base_id else f"auto-{_sha256(nf.name or '')[:8]}"

                s.run(
                    """
                    MERGE (ft:Feature {feature_id: $feature_id})
                    ON CREATE SET
                        ft.name = $name,
                        ft.actions = $actions,
                        ft.role = $role,
                        ft.created_at = $created_at,
                        ft.last_seen_at = $last_seen_at,
                        ft.is_prior = false
                    ON MATCH SET
                        ft.name = coalesce($name, ft.name),
                        ft.actions = coalesce($actions, ft.actions),
                        ft.role = coalesce($role, ft.role),
                        ft.last_seen_at = $last_seen_at,
                        ft.is_prior = coalesce(ft.is_prior, false)
                    """,
                    feature_id=auto_id,
                    name=nf.name,
                    actions=nf.actions,
                    role=nf.role,
                    created_at=now_ms,
                    last_seen_at=now_ms,
                )
                all_feature_ids.append(auto_id)

                # 同时把新发现的 feature 补充到内存中的 PRIOR_FEATURES，并尝试写回 prior_features.json
                try:
                    if PRIOR_FEATURES is not None:
                        # 避免重复添加同一个 feature_id
                        if not any(pf.id == auto_id for pf in PRIOR_FEATURES):
                            PRIOR_FEATURES.append(PriorFeature(
                                id=auto_id,
                                name=nf.name,
                                actions=nf.actions,
                                role=nf.role,
                            ))
                            if PRIOR_PATH:
                                save_prior_features(PRIOR_FEATURES, PRIOR_PATH)
                except Exception:
                    # JSON 写回失败时不影响 agent 运行
                    pass

            # 3) 把 Action 和所有相关 Feature 通过 ACHIEVE 连接起来（一个 action 多条边）
            for fid in all_feature_ids:
                s.run(
                    """
                    MATCH (a:Action {action_id: $action_id})
                    MATCH (f:Feature {feature_id: $feature_id})
                    MERGE (a)-[:ACHIEVE]->(f)
                    """,
                    action_id=action_id,
                    feature_id=fid,
                )

    except Exception:
        # 不要因为图写入失败把 agent 弄挂
        pass

def _attach_network_logging(page, global_apis: dict, active_apis_getter=None):
    """
    Attach request/response listeners that record into:
      - global_apis: dict (aggregated for the whole run)
      - active_apis_getter(): optional callable returning the current feature dict (per-feature)
    """
    def _ensure(map_obj, method, path):
        key = f"{method} {path}"
        rec = map_obj.setdefault(key, {"method": method, "path": path, "headers": [], "payloads": [], "statuses": set()})
        return rec

    def on_request(req):
        try:
            u = urlparse(req.url)
            method = req.method
            path = _normalize_path(u.path)
            # global
            rec = _ensure(global_apis, method, path)
            hdrs = {k.lower(): v for k, v in req.headers.items() if k.lower() in ("accept","content-type","authorization","x-requested-with")}
            if hdrs:
                rec["headers"].append(hdrs)
            try:
                body = req.post_data_json
            except Exception:
                body = req.post_data
            if body:
                if isinstance(body, str) and len(body) > 500:
                    body = body[:500]
                rec["payloads"].append(body)
            # per-feature
            if callable(active_apis_getter):
                feat_map = active_apis_getter()
                if isinstance(feat_map, dict):
                    rec2 = _ensure(feat_map, method, path)
                    if hdrs:
                        rec2["headers"].append(hdrs)
                    if body:
                        rec2["payloads"].append(body)
        except Exception:
            pass

    def on_response(res):
        try:
            u = urlparse(res.url)
            method = res.request.method
            path = _normalize_path(u.path)
            _ensure(global_apis, method, path)["statuses"].add(res.status)
            if callable(active_apis_getter):
                feat_map = active_apis_getter()
                if isinstance(feat_map, dict):
                    _ensure(feat_map, method, path)["statuses"].add(res.status)
        except Exception:
            pass

    page.on("request", on_request)
    page.on("response", on_response)

def _save_step_artifacts(page, base_dir: str, step_idx: int, tag: str = ""):
    os.makedirs(base_dir, exist_ok=True)
    prefix = os.path.join(base_dir, f"{step_idx:02d}_{tag}" if tag else f"{step_idx:02d}")
    # HTML
    try:
        html = page.content()
        with open(prefix + ".html", "w", encoding="utf-8") as f:
            f.write(html)
    except Exception:
        pass
    # Screenshot
    try:
        page.screenshot(path=prefix + ".png", full_page=True)
    except Exception:
        pass
    # Accessibility snapshot (JSON/YAML)
    try:
        acc = page.accessibility.snapshot()
        with open(prefix + ".aria.yml", "w", encoding="utf-8") as f:
            yaml.safe_dump(acc, f, allow_unicode=True, sort_keys=False)
    except Exception:
        pass
    # PDF (Chromium headless typically)
    try:
        page.pdf(path=prefix + ".pdf", print_background=True)
    except Exception:
        # ignore if not supported in non-headless mode
        pass

# Helper: Screenshot as base64 PNG (without writing to disk)
def _screenshot_b64(page) -> str:
    try:
        # get bytes directly to avoid writing to disk
        data = page.screenshot(type="png", full_page=True)
        return base64.b64encode(data).decode("utf-8")
    except Exception:
        return ""

def _query_clickables(page):
    """
    Collect candidate clickables with visible text from:
      1) ARIA roles: button/link/tab
      2) Generic interactive selectors: button, [role=button], a[href], [role=link], [role=tab], [tabindex], [onclick]
      3) Shadow DOM (pierce): :scope >>> <same selectors>
      4) All frames (main frame + iframes)
    Returns a list of {"name": str, "kind": "auto", "locator": Locator}
    """
    def _visible_name(el):
        # Prefer aria-label, then inner_text, then title, then textContent, then placeholder/alt/value, then href basename
        name = ""
        try:
            name = (el.get_attribute("aria-label") or "").strip()
            if name:
                return name
        except Exception:
            pass
        try:
            name = (el.inner_text(timeout=1000) or "").strip()
            if name:
                return name
        except Exception:
            pass
        try:
            name = (el.get_attribute("title") or "").strip()
            if name:
                return name
        except Exception:
            pass
        try:
            name = (el.text_content(timeout=1000) or "").strip()
            if name:
                return name
        except Exception:
            pass
        # Try to infer a name from SVG-based icon metadata (e.g., data-icon="folder")
        try:
            svg = el.locator("svg[data-icon]")
            if svg.count() > 0:
                icon_attr = (svg.nth(0).get_attribute("data-icon") or "").strip()
                if icon_attr:
                    return icon_attr
        except Exception:
            pass
        # New: try placeholder, alt, and value attributes as fallbacks
        for attr in ("placeholder", "alt", "value"):
            try:
                val = (el.get_attribute(attr) or "").strip()
                if val:
                    return val
            except Exception:
                continue
        try:
            href = el.get_attribute("href")
            if href:
                from urllib.parse import urlparse
                p = urlparse(href)
                base = (p.path.rsplit("/", 1)[-1] or p.path).strip()
                if base:
                    return base
        except Exception:
            pass
        return ""

    def _collect_from_scope(scope):
        candidates = []
        seen = set()

        # Strategy 1: roles
        role_locators = [
            ("button", scope.get_by_role("button")),
            ("link",   scope.get_by_role("link")),
            ("tab",    scope.get_by_role("tab")),
        ]
        for kind, loc in role_locators:
            try:
                cnt = loc.count()
            except Exception:
                cnt = 0
            for i in range(min(cnt, 400)):
                try:
                    el = loc.nth(i)
                    if not el.is_visible():
                        continue
                    name = _visible_name(el)
                    if not name:
                        continue
                    key = (kind, name)
                    if key in seen:
                        continue
                    seen.add(key)
                    candidates.append({"name": name, "kind": kind, "locator": el})
                except Exception:
                    continue

        # Strategy 2: generic CSS
        generic_sel = "button,[role=button],a[href],[role=link],[role=tab],[tabindex],[onclick]"
        loc = scope.locator(generic_sel)
        try:
            cnt = loc.count()
        except Exception:
            cnt = 0
        for i in range(min(cnt, 600)):
            try:
                el = loc.nth(i)
                if not el.is_visible():
                    continue
                name = _visible_name(el)
                if not name:
                    continue
                key = ("generic", name)
                if key in seen:
                    continue
                seen.add(key)
                candidates.append({"name": name, "kind": "generic", "locator": el})
            except Exception:
                continue

        # Strategy 3: shadow-piercing
        pierce_sel = ":scope >>> button, :scope >>> a[href], :scope >>> [role=button], :scope >>> [role=link], :scope >>> [role=tab]"
        loc = scope.locator(pierce_sel)
        try:
            cnt = loc.count()
        except Exception:
            cnt = 0
        for i in range(min(cnt, 400)):
            try:
                el = loc.nth(i)
                if not el.is_visible():
                    continue
                name = _visible_name(el)
                if not name:
                    continue
                key = ("shadow", name)
                if key in seen:
                    continue
                seen.add(key)
                candidates.append({"name": name, "kind": "shadow", "locator": el})
            except Exception:
                continue

        # Strategy 4: framework-specific icon-only sidebar buttons (e.g., el-custom sidebar triggers)
        # This is to catch clickable divs like:
        # <div class="el-custom-sidebar-tooltip__trigger ...">
        #   <div class="el-custom-sidebar-icon-container ..."><i><svg data-icon="folder">...</svg></i></div>
        # </div>
        try:
            special_sel = ".el-custom-sidebar-tooltip__trigger, .el-custom-sidebar-icon-container"
            loc = scope.locator(special_sel)
            try:
                cnt = loc.count()
            except Exception:
                cnt = 0
            for i in range(min(cnt, 200)):
                try:
                    el = loc.nth(i)
                    if not el.is_visible():
                        continue
                    name = _visible_name(el)
                    # If still no name, fall back to SVG icon metadata or a generic label
                    if not name:
                        try:
                            svg = el.locator("svg[data-icon]")
                            if svg.count() > 0:
                                icon_attr = (svg.nth(0).get_attribute("data-icon") or "").strip()
                                if icon_attr:
                                    name = icon_attr
                        except Exception:
                            pass
                    if not name:
                        name = "sidebar-icon"
                    key = ("sidebar", name)
                    if key in seen:
                        continue
                    seen.add(key)
                    candidates.append({"name": name, "kind": "sidebar", "locator": el})
                except Exception:
                    continue
        except Exception:
            pass

        return candidates

    # Collect from main page and all frames
    all_candidates = []
    try:
        all_candidates.extend(_collect_from_scope(page))
    except Exception:
        pass
    for fr in page.frames:
        try:
            all_candidates.extend(_collect_from_scope(fr))
        except Exception:
            continue

    # De-duplicate by name, keep first occurrence
    deduped = []
    names_seen = set()
    for c in all_candidates:
        nm = c["name"].strip()
        if not nm:
            continue
        if nm.lower() in names_seen:
            continue
        names_seen.add(nm.lower())
        deduped.append(c)

    # Debug: if empty, log a hint
    if not deduped:
        try:
            print(f"[agent] No clickables found. URL={page.url}. Possible reasons: login page, heavy shadow DOM, or content inside iframe requiring user interaction.")
        except Exception:
            pass
    return deduped


@retrying
def llm_choose_next_action(feature_goal: dict, url: str, clickables: List[dict]) -> dict:
    """
    Ask LLM to choose the best next clickable (by visible text) or say 'back'/'stop'.
    """
    prompt = (
        "You are navigating a web app to explore ONE feature end-to-end.\n"
        f"Current URL: {url}\n"
        "Given the current clickable elements (names), pick next action to make progress towards the feature.\n"
        "Return one of: {\"action\":\"click\",\"target\":\"<visible name>\"} or {\"action\":\"back\"} or {\"action\":\"stop\"}.\n"
        "Avoid destructive actions like 'Scan' or 'Settings'. Favor tabs, links, filters, 'Reports', 'History', 'Vulnerabilities', 'Components', 'SBOM'."
    )
    options = sorted({c["name"] for c in clickables})
    resp = client.responses.parse(
        model=MODEL_FEATURE,
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": prompt},
                {"type": "input_text", "text": "Feature goal:\n" + json.dumps(feature_goal, ensure_ascii=False)},
                {"type": "input_text", "text": "Clickables:\n" + json.dumps(options, ensure_ascii=False)}
            ]
        }],
        text_format=ActionChoice,
        timeout=REQ_TIMEOUT
    )
    parsed: ActionChoice = resp.output_parsed
    choice = {"action": parsed.action, "target": parsed.target}
    # Safety: avoid unsafe buttons
    if choice.get("action") == "click" and choice.get("target") in {"Scan", "Settings"}:
        choice = {"action": "stop"}
    return choice

# -------- Context-aware next action (HTML + ARIA) --------

# -------- Context-aware next action (Vision + ARIA, no URL/HTML) --------
@retrying

# Minimal natural-language description for a structured feature goal
def _feature_goal_to_text(goal: Dict[str, Any]) -> str:
    """
    Simplify structured feature data into a minimal natural-language description.
    """
    name = (goal.get("name") or "").strip()
    category = (goal.get("category") or "").strip()
    parts = []
    if name:
        parts.append(f"Feature: {name}.")
    if category:
        parts.append(f"Category: {category}.")
    return " ".join(parts)

def llm_choose_next_action_with_vision(feature_goal: dict, clickables: List[dict], aria_yaml_text: str, screenshot_b64: str) -> dict:
    """
    Decide next action using CURRENT PAGE CONTEXT:
    - Screenshot (vision) + ARIA snapshot text + list of visible clickable names.
    This avoids relying on URL/HTML which may be misleading after login.
    """
    aria_excerpt = aria_yaml_text[:20000] if aria_yaml_text else ""
    options = sorted({c["name"] for c in clickables})
    prompt = (
        "You are exploring ONE user-facing feature in a web app.\n"
        "Use the screenshot (visual layout) and ARIA tree text to understand the current page.\n"
        "From the provided clickable element NAMES, decide what to do next for this feature.\n"
        "- If clicking a candidate would likely move the user closer to using or fully understanding this feature, choose a click action.\n"
        "- If the current page already exposes everything relevant for this feature and further clicks would not add new behavior or views, choose stop.\n"
        "Prefer forward actions (tabs, section links, list rows, details, filters) over doing nothing, as long as they look potentially related to the feature.\n"
        "Avoid destructive actions like 'Scan' or 'Settings'.\n"
        "Return exactly one of:\n"
        '  {\"action\":\"click\",\"target\":\"<one of the provided names>\"}\n'
        '  {\"action\":\"stop\"}\n'
        "- If the screenshot shows a text input, search box, or form field relevant to this feature, you may return:\n"
        "    {\"action\":\"fill\",\"target\":\"<field name>\",\"value\":\"<text to input>\"}\n"
        "  The value should be short and generic unless the field context suggests a specific query (e.g., 'project', 'CVE-2023').\n"
    )
    # Build request
    content = [
        {"type": "input_text", "text": prompt},
        {"type": "input_text", "text": _feature_goal_to_text(feature_goal)},
        {"type": "input_text", "text": "Clickables:\n" + json.dumps(options, ensure_ascii=False)},
        {"type": "input_text", "text": "[ARIA EXCERPT]\n" + aria_excerpt},
    ]
    if screenshot_b64:
        content.insert(1, {"type": "input_image", "image_url": f"data:image/png;base64,{screenshot_b64}"})
    resp = client.responses.parse(
        model=MODEL_VISION,
        input=[{"role": "user", "content": content}],
        text_format=ActionChoice,
        timeout=REQ_TIMEOUT
    )
    parsed: ActionChoice = resp.output_parsed
    choice = {"action": parsed.action, "target": parsed.target}
    # Validate allowed actions
    act = choice.get("action")
    if act == "click" and (choice.get("target") in {"Scan", "Settings"}):
        choice = {"action": "stop"}
    if act not in {"click","stop","fill"}:
        choice = {"action":"stop"}
    return choice

# -------- Rank clickables for fallback --------

# -------- Rank clickables for fallback (Vision + ARIA, no URL/HTML) --------
@retrying
def llm_rank_clickables(feature_goal: dict, clickables: List[dict], aria_yaml_text: str, screenshot_b64: str, top_k: int = 3) -> List[str]:
    prompt = (
        "Rank clickable element names that most likely advance the given feature from the current page.\n"
        "Base your judgment on the screenshot and ARIA tree. Prefer tabs, details links, report/history, filters, and domain-relevant keywords.\n"
        "Avoid destructive actions (Scan/Settings). Return a short JSON list of names (subset of provided)."
    )
    options = sorted({c["name"] for c in clickables})
    content = [
        {"type": "input_text", "text": prompt},
        {"type": "input_text", "text": "Feature goal:\n" + json.dumps(feature_goal, ensure_ascii=False)},
        {"type": "input_text", "text": "Clickables:\n" + json.dumps(options, ensure_ascii=False)},
        {"type": "input_text", "text": "[ARIA EXCERPT]\n" + (aria_yaml_text[:10000] if aria_yaml_text else "")},
    ]
    if screenshot_b64:
        content.insert(1, {"type": "input_image", "image_url": f"data:image/png;base64,{screenshot_b64}"})
    resp = client.responses.parse(
        model=MODEL_VISION,
        input=[{"role": "user", "content": content}],
        text_format=list[str],
        timeout=REQ_TIMEOUT
    )
    ranked: List[str] = resp.output_parsed[:top_k]
    safe = [x for x in ranked if x not in {"Scan","Settings"}]
    return safe[:top_k]

def agent_explore_once(start_url: str, prior_path: str, out_dir: str, state_path: str | None = None):
    """
    One-round: use docs-driven prior, then explore with Playwright per feature.
    Each feature gets its own folder and step-by-step artifacts. Also export routes/apis files.
    """
    # Load prior features from JSON so agent and prior are fully decoupled
    if not prior_path or not os.path.exists(prior_path):
        raise FileNotFoundError(f"Prior features JSON not found: {prior_path}")
    prior_features: List[PriorFeature] = load_prior_features(prior_path)
    global PRIOR_FEATURES, PRIOR_PATH
    PRIOR_FEATURES = prior_features
    PRIOR_PATH = prior_path
    os.makedirs(out_dir, exist_ok=True)
    global_routes: set = set()
    global_apis: dict = {}


    # load login cookie
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        if state_path and os.path.exists(state_path):
            context = browser.new_context(storage_state=state_path)
        else:
            context = browser.new_context()
        page = context.new_page()
        # Connect to WM graph if live writing enabled
        if WRITE_LIVE_GRAPH:
            _wm_connect(
                os.environ.get("NEO4J_URI", args.neo4j_uri if 'args' in globals() else "bolt://localhost:7687"),
                os.environ.get("NEO4J_USER", "neo4j"),
                os.environ.get("NEO4J_PASSWORD", "12345678")
            )
        class _Holder: pass
        _apis_holder = _Holder()
        _apis_holder.current = None
        _attach_network_logging(page, global_apis, active_apis_getter=lambda: _apis_holder.current)

        # Go to start
        page.goto(start_url, wait_until="domcontentloaded")
        page.wait_for_timeout(AGENT_WAIT_NETWORK_IDLE_MS)
        _record_route(global_routes, page.url)
        # Write current page + discovered features to graph
        _wm_upsert_page_from_playwright(page)
        clickables_init = _query_clickables(page)
        # LLM gate: only keep user-facing features aligned with prior
        try:
            aria_obj0 = page.accessibility.snapshot()
            aria_text0 = yaml.safe_dump(aria_obj0, allow_unicode=True, sort_keys=False)
        except Exception:
            aria_text0 = ""
        sc0 = _screenshot_b64(page)
        # filtered0 = llm_filter_clickables_with_prior(clickables_init, feat, aria_text0, sc0)

        # For each feature to explorate
        for feat in prior_features:
            # Skip obviously unsafe features
            if feat.name in {"Scan", "Settings"}:
                continue

            feat_dir = os.path.join(out_dir, f"feature_{feat.id}")
            os.makedirs(feat_dir, exist_ok=True)

            # Per-feature route/API stores
            feature_routes: set = set()
            feature_apis: dict = {}
            _apis_holder.current = feature_apis
            last_action_id: str | None = None

            # feature header doc
            with open(os.path.join(feat_dir, "feature.md"), "w", encoding="utf-8") as fmd:
                fmd.write(f"# {feat.name}\n\n")
                fmd.write(f"- Role: {feat.role}\n- Actions: {', '.join(feat.actions)}\n")

            # Reset to start for each feature to ensure determinism
            try:
                page.goto(start_url, wait_until="domcontentloaded")
                page.wait_for_timeout(AGENT_WAIT_NETWORK_IDLE_MS)
            except PlaywrightTimeoutError:
                pass
            _save_step_artifacts(page, feat_dir, 0, "start")
            _record_route(global_routes, page.url)
            _record_route(feature_routes, page.url)
            # Write current page + discovered features to graph
            _wm_upsert_page_from_playwright(page)
            clickables_init = _query_clickables(page)
            try:
                aria_obj0 = page.accessibility.snapshot()
                aria_text0 = yaml.safe_dump(aria_obj0, allow_unicode=True, sort_keys=False)
            except Exception:
                aria_text0 = ""
            sc0 = _screenshot_b64(page)

            seen_urls = {page.url}
            # Track which (url, clickable_name) pairs have already been attempted for this feature
            tried_pairs: set[tuple[str, str]] = set()

            for step in range(1, AGENT_MAX_STEPS_PER_FEATURE + 1):
                _record_route(global_routes, page.url)
                _record_route(feature_routes, page.url)
                # Graph: upsert current page
                _wm_upsert_page_from_playwright(page)

                # Collect and LLM-filter clickables for the current feature on this page
                clickables_all = _query_clickables(page)
                try:
                    aria_obj_step = page.accessibility.snapshot()
                    aria_text_step = yaml.safe_dump(aria_obj_step, allow_unicode=True, sort_keys=False)
                except Exception:
                    aria_text_step = ""
                sc_step = _screenshot_b64(page)
                clickables = llm_filter_clickables_with_prior(clickables_all, feat, aria_text_step, sc_step)

                if not clickables:
                    # No relevant clickables for this feature on this page
                    _save_step_artifacts(page, feat_dir, step, "no-clickables")
                    break

                current_url = page.url
                # Pick the next clickable on this URL that we have not yet tried for this feature
                candidate = None
                for c in clickables:
                    name_norm = (c.get("name") or "").strip().lower()
                    if not name_norm:
                        continue
                    key = (current_url, name_norm)
                    if key not in tried_pairs:
                        candidate = c
                        break

                if candidate is None:
                    # All clickables on this page have been attempted for this feature; stop exploring further from here
                    _save_step_artifacts(page, feat_dir, step, "exhausted")
                    break

                # Mark this (url, clickable) as tried
                cand_name = (candidate.get("name") or "").strip()
                cand_name_norm = cand_name.lower()
                tried_pairs.add((current_url, cand_name_norm))

                # Prepare context for deciding what to do with THIS specific clickable
                try:
                    aria_obj = page.accessibility.snapshot()
                    aria_text = yaml.safe_dump(aria_obj, allow_unicode=True, sort_keys=False)
                except Exception:
                    aria_text = ""
                screenshot_b64 = _screenshot_b64(page)

                # Ask the model whether clicking this candidate is a useful step for this feature,
                # or whether the feature is already sufficiently explored on this page.
                choice = llm_choose_next_action_with_vision(asdict(feat), [candidate], aria_text, screenshot_b64)

                # 如果模型认为这个 clickable 对当前 feature 不需要（已经充分探索），就跳过这个 candidate
                # back（如果偶尔产生）也等价视为“跳过这个 clickable”
                act = choice.get("action")

                # stop/back → skip this clickable
                if act in {"stop","back"}:
                    continue

                # fill → perform form‑filling action
                if act == "fill":
                    field_name = choice.get("target") or ""
                    value = choice.get("value") or ""
                    locator = candidate.get("locator")
                    if locator is None:
                        continue
                    try:
                        locator.fill(value, timeout=AGENT_CLICK_TIMEOUT_MS)
                        page.wait_for_timeout(AGENT_WAIT_NETWORK_IDLE_MS)
                    except Exception:
                        continue

                    # record Action node
                    action_name = f"fill {field_name}"
                    action_desc = f"Fill '{field_name}' with '{value}' on page {current_url} for feature {feat.id}."
                    action_id = _wm_create_action_node(feat.id, current_url, "fill", action_name, action_desc)
                    _wm_link_action_next(last_action_id, action_id)
                    last_action_id = action_id
                    _wm_action_achieve_feature(action_id, feat.id)
                    # fill does not navigate, so continue on same page
                    continue

                # default required action: click
                if act != "click":
                    break

                # Execute the click on this candidate
                target = cand_name
                locator = candidate.get("locator")
                if locator is None:
                    # No valid locator; skip this clickable
                    continue

                src_url_for_edge = page.url
                try:
                    locator.click(timeout=AGENT_CLICK_TIMEOUT_MS)
                    page.wait_for_load_state("domcontentloaded", timeout=AGENT_CLICK_TIMEOUT_MS)
                    page.wait_for_timeout(AGENT_WAIT_NETWORK_IDLE_MS)
                except Exception:
                    # Click failed; go on to the next clickable on this page
                    continue

                _save_step_artifacts(page, feat_dir, step, "click")

                # Create Action node: click target on this page
                action_name = f"click {target}"
                action_desc = f"Click '{target}' on page {src_url_for_edge} for feature {feat.id}."
                action_id = _wm_create_action_node(feat.id, src_url_for_edge, "click", action_name, action_desc)
                _wm_link_action_next(last_action_id, action_id)
                last_action_id = action_id

                _wm_feature_to_page(src_url_for_edge, {"name": target, "kind": "auto"}, page.url)
                _wm_page_to_page(src_url_for_edge, page.url, via="click", anchor_text=target)
                if page.url != src_url_for_edge:
                    # 每次 click 后，用 GPT 判断这个 Action 实现了哪些 prior / 新 feature
                    _wm_action_achieve_feature(action_id, feat.id)
                    _wm_action_forward_page(action_id, page.url)

                # If clicking did not change the URL, stay on the same page and continue trying other clickables
                if page.url == src_url_for_edge:
                    continue

                # If we have already seen this URL for this feature, avoid looping endlessly
                if page.url in seen_urls:
                    _save_step_artifacts(page, feat_dir, step, "loop-detected")
                    break

                # New URL discovered for this feature: remember it and continue exploring from the new page
                seen_urls.add(page.url)

            # Each feature now has its own routes.json and apis.json for debugging and documentation.
            # Persist per-feature routes/apis
            with open(os.path.join(feat_dir, "routes.json"), "w", encoding="utf-8") as f:
                json.dump({"routes": sorted(feature_routes)}, f, ensure_ascii=False, indent=2)
            feat_apis_serializable = []
            for v in feature_apis.values():
                feat_apis_serializable.append({
                    **{k: v[k] for k in ("method","path")},
                    "headers": v["headers"],
                    "payloads": v["payloads"],
                    "statuses": sorted(list(v["statuses"]))
                })
            with open(os.path.join(feat_dir, "apis.json"), "w", encoding="utf-8") as f:
                json.dump({"apis": feat_apis_serializable}, f, ensure_ascii=False, indent=2)
            # Record API calls for this feature's final page
            _wm_calls_api_from_page(page.url, feature_apis)

        # Save routes/apis
        routes_path = os.path.join(out_dir, "routes.json")
        apis_path = os.path.join(out_dir, "apis.json")
        with open(routes_path, "w", encoding="utf-8") as f:
            json.dump({"routes": sorted(global_routes)}, f, ensure_ascii=False, indent=2)
        apis_serializable = []
        for v in global_apis.values():
            apis_serializable.append({
                **{k: v[k] for k in ("method","path")},
                "headers": v["headers"],
                "payloads": v["payloads"],
                "statuses": sorted(list(v["statuses"]))
            })
        with open(apis_path, "w", encoding="utf-8") as f:
            json.dump({"apis": apis_serializable}, f, ensure_ascii=False, indent=2)
        # Record global API calls
        _wm_calls_api_from_page(page.url, global_apis)

        browser.close()

@retrying
def llm_extract_features_from_html(html_chunk: str) -> List[Feature]:
    """Extract FEATURES only from HTML (no screenshots)."""
    prompt = (
        "You are a senior UX analyst. From ONLY the given HTML markup, extract end-user facing FEATURES "
        "(navigation items, pages, cards, tables, search/sort/filter controls, buttons like 'Scan', dialogs, "
        "notifications, pagination, language/org/user menus). Do not invent data. Prefer names visible in DOM text.\n\n"
        "For each feature, provide: id (kebab-case), name, category, actions, page, short evidence snippet, "
        "optional CSS-like locator, and confidence 1..5.\n"
        "Treat dashboard cards and list pages as separate features. Merge obvious duplicates within the same chunk."
    )
    resp = client.responses.parse(
        model=MODEL_FEATURE,
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": prompt},
                {"type": "input_text", "text": html_chunk}
            ]
        }],
        text_format=FeaturesSchema,
        timeout=REQ_TIMEOUT
    )
    data: FeaturesSchema = resp.output_parsed
    out: List[Feature] = []
    for f in data.features:
        out.append(Feature(
            id=f.id, name=f.name, category=f.category, actions=f.actions,
            page=f.page, evidence=f.evidence, source="html",
            confidence=f.confidence, locator=f.locator or ""
        ))
    return out


@retrying
def llm_extract_features_from_image(image_b64: str, mime: str, filename: str) -> List[Feature]:
    prompt = (
        "From ONLY this screenshot, list the visible user-facing FEATURES: "
        "left nav items, top menus (notifications/language/org/user), dashboard cards, charts, list rows, "
        "buttons (e.g., Scan), search/sort/filter controls, pagination hints (e.g., '15 of 913 Projects'), "
        "expand/collapse arrows, and obvious links (e.g., project names, CVE). "
        "Do not hallucinate; rely on visible labels. Guess the 'page' name from visible context.\n"
        f"Screenshot filename: {filename}"
    )
    resp = client.responses.parse(
        model=MODEL_VISION,
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": prompt},
                {"type": "input_image", "image_url": f"data:{mime};base64,{image_b64}"}
            ]
        }],
        text_format=FeaturesSchema,
        timeout=REQ_TIMEOUT
    )
    data: FeaturesSchema = resp.output_parsed
    out: List[Feature] = []
    for f in data.features:
        out.append(Feature(
            id=f.id, name=f.name, category=f.category, actions=f.actions,
            page=f.page, evidence=f.evidence, source="image",
            confidence=f.confidence, locator=f.locator or ""
        ))
    return out


@retrying
def llm_reconcile_features(html_features: List[Dict[str, Any]], img_features: List[Dict[str, Any]]) -> List[ReconciledFeature]:
    """Merge & cross-check HTML vs Screenshot features."""
    instructions = (
        "Reconcile two feature lists from (A) HTML-only and (B) screenshot-only.\n"
        "- Merge duplicates and near-duplicates; prefer screenshot naming for visible user labels.\n"
        "- status='both' if matched in A and B; otherwise 'html_only' or 'image_only'.\n"
        "- Keep a compact evidence string, keep best locator, actions is the union.\n"
        "- Ensure stable, kebab-case ids."
    )
    resp = client.responses.parse(
        model=MODEL_FEATURE,
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": instructions},
                {"type": "input_text", "text": "HTML features:\n" + json.dumps(html_features, ensure_ascii=False)},
                {"type": "input_text", "text": "Screenshot features:\n" + json.dumps(img_features, ensure_ascii=False)}
            ]
        }],
        text_format=ReconciledSchema,
        timeout=REQ_TIMEOUT
    )
    data: ReconciledSchema = resp.output_parsed
    out: List[ReconciledFeature] = []
    for f in data.features:
        out.append(ReconciledFeature(
            id=f.id, name=f.name, category=f.category, actions=f.actions,
            page=f.page, evidence=f.evidence, source="merged",
            confidence=f.confidence, locator=f.locator or "",
            status=f.status, matched_with=f.matched_with if f.matched_with is not None else []
        ))
    return out


@retrying
def llm_generate_user_stories(features: List[Dict[str, Any]]) -> List[UserStory]:
    """Generate User Stories (AC + Priority) from reconciled features with semantic dedup/merge."""
    prompt = (
        "You are a product owner. Convert the reconciled feature list into concise USER STORIES.\n"
        "Rules:\n"
        "1) 'As a <actor>, I want to <goal>, so that <benefit>.'\n"
        "2) Title is Verb + Noun (e.g., 'Trigger Scan', 'Search Projects').\n"
        "3) Provide 2–4 Acceptance Criteria per story, testable and UI-oriented.\n"
        "4) Deduplicate: merge CRUD siblings into 'Manage X' when appropriate.\n"
        "5) Prioritize: High (core navigation/scan/search), Medium (dashboards/logs), Low (i18n/help).\n"
        "6) Keep related_features as the list of feature ids supporting the story."
    )
    resp = client.responses.parse(
        model=MODEL_FEATURE,
        input=[{
            "role": "user",
            "content": [
                {"type": "input_text", "text": prompt},
                {"type": "input_text", "text": json.dumps(features, ensure_ascii=False)}
            ]
        }],
        text_format=UserStoriesSchema,
        timeout=REQ_TIMEOUT
    )
    data: UserStoriesSchema = resp.output_parsed
    out: List[UserStory] = []
    for s in data.stories:
        out.append(UserStory(
            id=s.id, actor=s.actor, title=s.title, story=s.story,
            acceptance_criteria=s.acceptance_criteria, related_features=s.related_features,
            priority=s.priority
        ))
    return out

# --------------- Feature Description Summarization ---------------

def _feature_goal_to_text(goal: Dict[str, Any]) -> str:
    """Convert structured prior feature info into a short natural-language summary for the agent."""
    name = (goal.get("name") or "").strip()
    role = (goal.get("role") or "").strip()
    actions = goal.get("actions") or []
    actions_str = ", ".join(actions) if isinstance(actions, list) else str(actions)
    parts = []
    if name:
        parts.append(f"Feature: {name}.")
    if role:
        parts.append(f"Role: {role}.")
    if actions_str:
        parts.append(f"Typical actions: {actions_str}.")
    base = " ".join(parts) or "Feature."
    try:
        summary_prompt = (
            f"Summarize the following feature description into one or two natural English sentences "
            f"explaining what the system likely does and how the UI might support it:\n\n{base}\n\n"
            "Example output: 'A user can manage projects by searching, filtering and opening project details from the UI.'"
        )
        resp = client.responses.create(
            model="gpt-4o-mini",
            input=[{"role": "user", "content": summary_prompt}],
            temperature=0.4,
        )
        summary_text = (resp.output_text or "").strip()
    except Exception:
        summary_text = ""
    if summary_text:
        return f"{base}\n{summary_text}"
    return base

# --------------- CLI -----------------


# Helper to allow directories for --html
def _normalize_html_inputs(html_args: List[str]) -> List[str]:
    paths = []
    for p in html_args:
        if os.path.isdir(p):
            for root, _, files in os.walk(p):
                for fn in files:
                    if fn.lower().endswith((".html", ".htm")):
                        paths.append(os.path.join(root, fn))
        else:
            paths.append(p)
    return sorted(set(paths))

# --- Helper: slugify page names for synthetic Page URLs ---
import unicodedata
import string

def _slugify(text: str) -> str:
    s = unicodedata.normalize("NFKD", text or "").strip().lower()
    allowed = string.ascii_lowercase + string.digits + "-"
    s = s.replace(" ", "-")
    s = "".join(ch if ch in allowed else "-" for ch in s)
    while "--" in s:
        s = s.replace("--", "-")
    return s.strip("-") or "page"


# Helper to clear all nodes and relationships in Neo4j
def clear_database(driver):
    """清空数据库中的所有节点和关系"""
    with driver.session() as session:
        session.run("MATCH ()-[r]->() DELETE r")
        session.run("MATCH (n) DELETE n")
    print("✅ 数据库已清空")

def _mk_feature_id_local(page_url: str, selector: str) -> str:
    return f"{page_url}|{selector}"

def build_working_memory_from_prior(prior_path, neo4j_uri, neo4j_user, neo4j_password):
    prior = load_prior_features(prior_path)
    features_payload = []
    ts = int(datetime.now(timezone.utc).timestamp() * 1000)
    for feat in prior:
        features_payload.append({
            "feature_id": feat.id,
            "name": feat.name,
            "actions": feat.actions,
            "role": feat.role,
            "created_at": ts,
            "last_seen_at": ts,
        })

    imp = WebsiteWMImporter(neo4j_uri, neo4j_user, neo4j_password)
    try:
        imp.connect()
        # Clear the entire database (all nodes and relationships)
        clear_database(imp._graph._driver)
        imp.ensure_schema()
        # Insert Feature nodes ONLY
        with imp._graph._driver.session() as s:
            for f in features_payload:
                s.run(
                    """
                    MERGE (ft:Feature {feature_id: $feature_id})
                    ON CREATE SET
                        ft.name = $name,
                        ft.actions = $actions,
                        ft.role = $role,
                        ft.created_at = $created_at,
                        ft.last_seen_at = $last_seen_at,
                        ft.is_prior = true
                    ON MATCH SET
                        ft.name = coalesce($name, ft.name),
                        ft.actions = coalesce($actions, ft.actions),
                        ft.role = coalesce($role, ft.role),
                        ft.last_seen_at = $last_seen_at,
                        ft.is_prior = true
                    """,
                    feature_id=f["feature_id"],
                    name=f["name"],
                    actions=f["actions"],
                    role=f["role"],
                    created_at=f["created_at"],
                    last_seen_at=f["last_seen_at"],
                )
    finally:
        imp.close()

# --------------- Export -----------------

def save_features_excel(features: List[ReconciledFeature], path: str):
    wb = Workbook()
    ws = wb.active
    headers = ["ID", "Name", "Category", "Actions", "Page", "Status", "Confidence", "Locator", "Evidence"]
    ws.append(headers)
    for f in features:
        ws.append([
            f.id, f.name, f.category, ", ".join(sorted(set(f.actions))), f.page,
            f.status, f.confidence, f.locator, f.evidence
        ])
    for col in "ABCDEFGHI":
        ws.column_dimensions[col].width = 18 if col != "I" else 40
    for row in ws.iter_rows(min_row=2, min_col=9, max_col=9):
        for cell in row:
            cell.alignment = Alignment(wrapText=True)
    wb.save(path)


def save_stories_excel(stories: List[UserStory], path: str):
    wb = Workbook()
    ws = wb.active
    headers = ["ID", "Actor", "Title", "Story", "Acceptance Criteria", "Related Features", "Priority"]
    ws.append(headers)
    for s in stories:
        ws.append([
            s.id, s.actor, s.title, s.story,
            "\n".join(s.acceptance_criteria),
            ", ".join(s.related_features),
            s.priority
        ])
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 10
    for col in "DE":
        for row in ws.iter_rows(min_row=2, min_col=ord(col)-64, max_col=ord(col)-64):
            for cell in row:
                cell.alignment = Alignment(wrapText=True)
    wb.save(path)


# --------------- Pipeline -----------------

def run_pipeline(html_files: List[str], screenshot_dir: str, out_dir: str):
    os.makedirs(out_dir, exist_ok=True)

    # Step 1: HTML → features
    print(f"[1/4] HTML → features: scanning {len(html_files)} HTML file(s)...")
    html_features: List[Feature] = []
    for hp in html_files:
        print(f"    - Reading {hp}")
        chunks = read_html_as_chunks(hp)
        for ch in chunks:
            html_features.extend(llm_extract_features_from_html(ch))

    # Simple dedup (same page+name, keep highest confidence)
    before_dedup = len(html_features)
    dedup: Dict[Any, Feature] = {}
    for f in html_features:
        key = (f.page.lower().strip(), f.name.lower().strip())
        if key not in dedup or dedup[key].confidence < f.confidence:
            dedup[key] = f
    html_features = list(dedup.values())
    after_dedup = len(html_features)
    print(f"    - Deduplicated HTML features: {before_dedup} → {after_dedup}")

    # Step 2: Screenshots → features
    print(f"[2/4] Screenshots → features: scanning directory '{screenshot_dir}' ...")
    img_features: List[Feature] = []
    img_filenames = []
    if os.path.isdir(screenshot_dir):
        for fn in sorted(os.listdir(screenshot_dir)):
            if fn.lower().endswith((".png", ".jpg", ".jpeg", ".webp")):
                img_filenames.append(fn)
    print(f"    - Found {len(img_filenames)} image file(s).")
    for fn in img_filenames:
        print(f"    - Analyzing {fn}")
        b64, mime = image_to_base64(os.path.join(screenshot_dir, fn))
        img_features.extend(llm_extract_features_from_image(b64, mime, fn))
    print(f"    - Extracted {len(img_features)} image features.")

    # Step 3: Cross-check & reconcile
    print(f"[3/4] Reconciling features (html={len(html_features)}, images={len(img_features)}) ...")
    html_payload = [asdict(f) for f in html_features]
    img_payload  = [asdict(f) for f in img_features]
    reconciled: List[ReconciledFeature] = llm_reconcile_features(html_payload, img_payload)
    both = sum(1 for x in reconciled if x.status == "both")
    html_only = sum(1 for x in reconciled if x.status == "html_only")
    image_only = sum(1 for x in reconciled if x.status == "image_only")
    print(f"    - Reconciled features: {len(reconciled)} (both={both}, html_only={html_only}, image_only={image_only})")

    # Step 4: User Stories（含语义合并）
    print(f"[4/4] Generating User Stories from {len(reconciled)} reconciled features ...")
    stories: List[UserStory] = llm_generate_user_stories([asdict(f) for f in reconciled])
    print(f"    - Generated {len(stories)} user stories.")

    # Save outputs
    print("[Saving] Writing JSON and Excel outputs ...")
    os.makedirs(out_dir, exist_ok=True)
    json_path = os.path.join(out_dir, "reconciled_features.json")
    features_xlsx_path = os.path.join(out_dir, "Web_Features.xlsx")
    stories_xlsx_path = os.path.join(out_dir, "User_Stories.xlsx")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump([asdict(x) for x in reconciled], f, ensure_ascii=False, indent=2)

    save_features_excel(reconciled, features_xlsx_path)
    save_stories_excel(stories, stories_xlsx_path)

    # Console summary
    print(f"[OK] Reconciled features: {len(reconciled)} (both={both}, html_only={html_only}, image_only={image_only})")
    print(f"[OK] Stories: {len(stories)}")
    print(f"Saved to: {out_dir}")
    print(f"  - JSON: {json_path}")
    print(f"  - Features Excel: {features_xlsx_path}")
    print(f"  - Stories Excel: {stories_xlsx_path}")


# --------------- CLI -----------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Reverse Website → Features → User Stories (HTML + Screenshots) OR One-Round Agent Explore (agent reads prior from out/prior_features.json)"
    )
    parser.add_argument("--mode", choices=["pipeline","agent"], default="agent", help="pipeline: offline HTML+shots; agent: one-round docs→explore")
    parser.add_argument("--start", default="https://v4staging.scantist.io", help="Start URL for agent mode")
    parser.add_argument("--docs", default="https://docs.scantist.io/en-US/", help="Comma-separated docs URLs for prior features")
    parser.add_argument("--out", default="./web/Record/memory2", help="Output directory (also per-feature docs in agent mode)")
    parser.add_argument("--state", default="./web/Record/state.json", help="Path to Playwright storage state JSON (logged-in session). If provided, the agent will reuse this login state.")
    # legacy pipeline args
    parser.add_argument("--html", nargs="+", default=["./web/Scantist.html"], help="HTML files or directories (pipeline mode)")
    parser.add_argument("--shots", default='./web/', help="Directory containing screenshots (pipeline mode)")
    # Neo4j/graph export options
    parser.add_argument("--neo4j_uri", default="bolt://localhost:7687")
    parser.add_argument("--neo4j_user", default="neo4j")
    parser.add_argument("--neo4j_password", default="12345678")
    parser.add_argument("--write_graph", default=True, help="After producing prior features, write Page/Feature nodes to Neo4j (node-only graph)")
    args = parser.parse_args()

    if args.mode == "pipeline":
        html_files = _normalize_html_inputs(args.html)
        if not html_files:
            raise FileNotFoundError("No HTML files found. Pass .html/.htm files or a directory containing them.")
        run_pipeline(html_files, args.shots, args.out)
    else:
        # one-round agent mode with decoupled prior
        os.makedirs(args.out, exist_ok=True)
        prior_path = os.path.join(args.out, "prior_features.json")
        if os.path.exists(prior_path):
            print(f"[Agent] Using existing prior at {prior_path}")
        else:
            doc_urls = [u.strip() for u in (args.docs or "").split(",") if u.strip()]
            docs_text = crawl_docs_text(doc_urls)
            prior = llm_prior_features_from_docs(docs_text)
            save_prior_features(prior, prior_path)
            print(f"[Agent] Wrote prior to {prior_path} (features={len(prior)})")
        # Write prior-only graph with feature nodes to Neo4j
        if args.write_graph:
            if not (args.neo4j_uri and args.neo4j_user and args.neo4j_password):
                raise RuntimeError("--write-graph requires --neo4j-uri, --neo4j-user, and --neo4j-password (or set NEO4J_* env vars)")
            build_working_memory_from_prior(prior_path, args.neo4j_uri, args.neo4j_user, args.neo4j_password)
            print("[Agent] Wrote prior-only working-memory graph (feature nodes).")
        # Now explore once, per feature → per doc
        agent_explore_once(args.start, prior_path, args.out, args.state or None)
        _wm_close()
        print(f"[OK] Agent one-round complete. Outputs in: {args.out}")
